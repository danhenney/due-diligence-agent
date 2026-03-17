"""Document preprocessor — converts uploaded PDF/Excel to MD before agent execution.

Runs as Step 0.5 in the pipeline. Extracts full text from all uploaded documents,
splits large PDFs by section, converts Excel to MD tables, and assigns docs to agents
by keyword matching.

This ensures agents read the FULL content of all documents, not just first 50 pages.
"""
from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any


# Agent-to-keyword mapping for automatic doc assignment
AGENT_KEYWORDS: dict[str, list[str]] = {
    "financial_analysis": [
        "재무", "손익", "매출", "revenue", "financial", "P&L", "대차대조",
        "balance", "현금흐름", "cash flow", "감사", "audit", "투자",
        "investment", "valuation", "프로젝션", "projection", "forecast",
    ],
    "market_analysis": [
        "시장", "market", "TAM", "SAM", "산업", "industry", "트렌드",
        "trend", "성장", "growth", "규모",
    ],
    "tech_analysis": [
        "기술", "tech", "R&D", "특허", "patent", "IP", "아키텍처",
        "architecture", "API", "model", "플랫폼", "platform",
    ],
    "competitor_analysis": [
        "경쟁", "competitor", "비교", "benchmark", "vs", "점유율", "share",
    ],
    "legal_regulatory": [
        "법률", "legal", "규제", "regulatory", "소송", "컴플라이언스",
        "compliance", "ESG", "지배구조", "governance",
    ],
    "team_analysis": [
        "팀", "team", "경영진", "CEO", "CTO", "조직", "organization",
        "인사", "HR", "임원",
    ],
}


def preprocess_documents(
    uploaded_docs: list[str],
    output_dir: str,
) -> dict[str, list[str]]:
    """Preprocess all uploaded docs into MD files and assign to agents.

    Args:
        uploaded_docs: List of file paths (PDF, Excel).
        output_dir:    Directory to write preprocessed MD files.

    Returns:
        Dict mapping agent_name -> list of preprocessed MD file paths.
    """
    os.makedirs(output_dir, exist_ok=True)
    all_entries: list[dict[str, Any]] = []

    for doc_path in uploaded_docs:
        if not os.path.exists(doc_path):
            continue

        ext = Path(doc_path).suffix.lower()
        base_name = Path(doc_path).stem

        if ext == ".pdf":
            entries = _process_pdf(doc_path, base_name, output_dir)
        elif ext in (".xlsx", ".xls"):
            entries = _process_excel(doc_path, base_name, output_dir)
        else:
            continue

        all_entries.extend(entries)

    # Assign to agents by keyword matching
    assignments: dict[str, list[str]] = {agent: [] for agent in AGENT_KEYWORDS}
    for entry in all_entries:
        text_to_match = (entry["file"] + " " + entry.get("section", "") + " " + entry["source"]).lower()
        matched = False
        for agent, keywords in AGENT_KEYWORDS.items():
            if any(kw.lower() in text_to_match for kw in keywords):
                assignments[agent].append(entry["path"])
                matched = True
        if not matched:
            # Unmatched docs go to all agents
            for agent in assignments:
                assignments[agent].append(entry["path"])

    # Write index file
    _write_index(all_entries, assignments, output_dir)

    return assignments


def _process_pdf(doc_path: str, base_name: str, output_dir: str) -> list[dict]:
    """Extract PDF text and split large docs by section."""
    try:
        import fitz
    except ImportError:
        return []

    entries = []
    doc = fitz.open(doc_path)
    pages = len(doc)
    full_text = ""
    for i in range(pages):
        full_text += f"\n--- Page {i + 1} ---\n"
        full_text += doc[i].get_text()

    if pages > 30:
        # Split by headings
        sections = _split_by_headings(full_text)
        for idx, sec in enumerate(sections):
            sec_name = re.sub(r"[^\w가-힣\s]", "", sec["title"])[:40].strip() or f"section_{idx}"
            sec_filename = f"{base_name}_{idx:02d}_{sec_name}.md"
            sec_path = os.path.join(output_dir, sec_filename)
            with open(sec_path, "w", encoding="utf-8") as f:
                f.write(f"# {sec['title']}\n\nSource: {os.path.basename(doc_path)} ({pages}p)\n\n")
                f.write(sec["content"])
            entries.append({
                "file": sec_filename,
                "path": sec_path,
                "source": os.path.basename(doc_path),
                "type": "pdf_section",
                "section": sec["title"],
                "chars": len(sec["content"]),
            })
    else:
        md_filename = f"{base_name}.md"
        md_path = os.path.join(output_dir, md_filename)
        with open(md_path, "w", encoding="utf-8") as f:
            f.write(f"# {os.path.basename(doc_path)}\n\nPages: {pages}\n\n")
            f.write(full_text)
        entries.append({
            "file": md_filename,
            "path": md_path,
            "source": os.path.basename(doc_path),
            "type": "pdf_full",
            "pages": pages,
            "chars": len(full_text),
        })

    return entries


def _process_excel(doc_path: str, base_name: str, output_dir: str) -> list[dict]:
    """Convert Excel to MD with full data (not sampled)."""
    try:
        import openpyxl
    except ImportError:
        return []

    wb = openpyxl.load_workbook(doc_path, data_only=True, read_only=True)
    md_content = f"# {os.path.basename(doc_path)}\n\n"

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        md_content += f"## Sheet: {sn} ({len(rows)} rows x {ws.max_column} cols)\n\n"

        if len(rows) <= 500:
            # Full table output
            if rows:
                header = rows[0]
                md_content += "| " + " | ".join([str(c)[:80] if c else "" for c in header]) + " |\n"
                md_content += "|" + "---|" * len(header) + "\n"
                for row in rows[1:]:
                    vals = [str(c)[:80] if c is not None else "" for c in row]
                    md_content += "| " + " | ".join(vals) + " |\n"
        else:
            # Large sheet: TSV format for compactness
            md_content += f"(Large sheet: {len(rows)} rows — full TSV output)\n\n```tsv\n"
            for row in rows:
                vals = [str(c)[:60] if c is not None else "" for c in row]
                md_content += "\t".join(vals) + "\n"
            md_content += "```\n"
        md_content += "\n"

    md_filename = f"{base_name}.md"
    md_path = os.path.join(output_dir, md_filename)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)

    return [{
        "file": md_filename,
        "path": md_path,
        "source": os.path.basename(doc_path),
        "type": "excel",
        "sheets": len(wb.sheetnames),
        "chars": len(md_content),
    }]


def _split_by_headings(text: str) -> list[dict]:
    """Split text by detected heading patterns."""
    sections = []
    current = {"title": "서두", "content": ""}

    for line in text.split("\n"):
        stripped = line.strip()
        if (
            re.match(r"^(제?\s*\d+[장부절편]|[IVX]+\.|[A-Z][A-Z\s]{3,}$|\d+\.\s+[가-힣A-Z])", stripped)
            and len(stripped) < 100
        ):
            if current["content"].strip():
                sections.append(current)
            current = {"title": stripped[:80], "content": ""}
        current["content"] += line + "\n"

    if current["content"].strip():
        sections.append(current)

    return sections


def _write_index(
    entries: list[dict],
    assignments: dict[str, list[str]],
    output_dir: str,
) -> None:
    """Write _doc_index.md summarising all preprocessed files and assignments."""
    index_path = os.path.join(output_dir, "_doc_index.md")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(f"# Document Index ({len(entries)} preprocessed files)\n\n")
        f.write("## File List\n\n")
        for e in entries:
            f.write(f"- **{e['file']}** <- {e['source']} | {e['type']} | {e['chars']:,} chars\n")
        f.write("\n## Agent Assignments\n\n")
        for agent, paths in assignments.items():
            f.write(f"### {agent} ({len(paths)} files)\n")
            for p in paths:
                f.write(f"- {os.path.basename(p)}\n")
            f.write("\n")
